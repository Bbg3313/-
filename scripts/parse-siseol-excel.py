# -*- coding: utf-8 -*-
"""Parse 시술정리.xlsx (세로 블록) → scripts/siseol-parsed.json. UTF-8."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    raise

LABEL_TO_FIELD: dict[str, str] = {
    "시술정보": "procedureInfo",
    "마취시간": "anesthesiaTime",
    "시술시간": "procedureTime",
    "회복기간": "recoveryPeriod",
    "유지기간": "effectDuration",
    "재시술주기": "retreatmentInterval",
    "주의사항": "precautions",
}


def norm_label(s: str) -> str | None:
    if not s:
        return None
    t = s.strip()
    if t.startswith("주의사항"):
        return "주의사항"
    if t in LABEL_TO_FIELD:
        return t
    return None


def read_col_a(ws) -> list[str | None]:
    out: list[str | None] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 2000, max_col=1, values_only=True):
        out.append(row[0] if row else None)
    return out


def is_stop_boundary(line: str) -> bool:
    if not line:
        return False
    if line.startswith("▶"):
        return True
    if re.match(r"^\s*\[", line):
        return True
    return norm_label(line) is not None


def join_lines(lines: list[str]) -> str:
    return "\n\n".join(x for x in lines if x).strip()


def parse_block(lines: list[str | None]) -> dict:
    raw = [("" if x is None else str(x)).strip() for x in lines]
    while raw and not raw[0]:
        raw.pop(0)

    area = ""
    if raw and raw[0].startswith("*"):
        area = raw[0]
        raw = raw[1:]
        while raw and not raw[0]:
            raw.pop(0)

    fields = {k: "" for k in ("procedureInfo", "anesthesiaTime", "procedureTime", "recoveryPeriod", "effectDuration", "retreatmentInterval", "precautions")}
    i = 0

    # ---- 시술정보 (라벨 있으면 소비, 없으면 곧바로 본문) ----
    if i < len(raw) and norm_label(raw[i]) == "시술정보":
        i += 1
    proc: list[str] = []
    while i < len(raw):
        line = raw[i]
        if not line:
            i += 1
            continue
        if norm_label(line) in ("마취시간", "시술시간"):
            break
        if is_stop_boundary(line) and norm_label(line) is None:
            break
        proc.append(line)
        i += 1
    fields["procedureInfo"] = join_lines(proc)

    # ---- 라벨:값 순서 ----
    label_order = ["마취시간", "시술시간", "회복기간", "유지기간", "재시술주기", "주의사항"]
    for li, lab in enumerate(label_order):
        if i >= len(raw):
            break
        if not raw[i]:
            i += 1
            continue
        nl = norm_label(raw[i])
        if nl != lab:
            continue
        i += 1
        vals: list[str] = []
        next_labels = set(label_order[li + 1 :])
        while i < len(raw):
            line = raw[i]
            if not line:
                i += 1
                continue
            if line.startswith("▶"):
                break
            if re.match(r"^\s*\[", line):
                break
            n2 = norm_label(line)
            if n2 and n2 in next_labels:
                break
            vals.append(line)
            i += 1
        key = LABEL_TO_FIELD[lab]
        fields[key] = join_lines(vals)

    return {"area": area or "-", **fields}


def raw_text_until_next(cells: list[str | None], start: int, end: int) -> str:
    lines: list[str] = []
    for j in range(start, min(end, len(cells))):
        v = cells[j]
        if v is None:
            continue
        s = str(v).strip()
        if s:
            lines.append(s)
    return join_lines(lines)


def find_blocks(cells: list[str | None]) -> list[tuple[int, str, int]]:
    starts: list[tuple[int, str]] = []
    for idx, v in enumerate(cells):
        if not v or not isinstance(v, str):
            continue
        t = v.strip()
        if t.startswith("▶"):
            starts.append((idx, t.lstrip("▶").strip()))
    blocks: list[tuple[int, str, int]] = []
    for j, (idx, title) in enumerate(starts):
        end = starts[j + 1][0] if j + 1 < len(starts) else len(cells)
        blocks.append((idx, title, end))
    return blocks


def main() -> None:
    default = Path.home() / "Documents" / "카카오톡 받은 파일" / "시술정리.xlsx"
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else default
    if not path.is_file():
        print(f"Missing file: {path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cells = read_col_a(ws)
    wb.close()

    blocks = find_blocks(cells)
    out: list[dict] = []
    for idx, title, end in blocks:
        inner = cells[idx + 1 : end]
        data = parse_block(inner)
        out.append({"excelTitle": title, "startRow": idx + 1, **data})

    merge_into_violet = {"시술 추천법", "주기", "브이올렛 시술원리", "브이올렛 시술 위치"}
    merged: list[dict] = []
    bi = 0
    while bi < len(blocks):
        idx, title, end = blocks[bi]
        cur = out[bi]
        if title == "브이올렛":
            extras: list[str] = []
            bj = bi + 1
            while bj < len(blocks) and blocks[bj][1] in merge_into_violet:
                sidx, stitle, send = blocks[bj]
                raw = raw_text_until_next(cells, sidx + 1, send)
                if raw:
                    extras.append(f"【{stitle}】\n{raw}")
                bj += 1
            if extras:
                cur["procedureInfo"] = (cur["procedureInfo"] + "\n\n" + "\n\n".join(extras)).strip()
            merged.append(cur)
            bi = bj
            continue
        if title == "덴서티 팁 종류 설명" and merged:
            prev = merged[-1]
            raw = raw_text_until_next(cells, idx + 1, end)
            if raw:
                prev["procedureInfo"] = (prev["procedureInfo"] + "\n\n【덴서티 팁 종류】\n" + raw).strip()
            bi += 1
            continue
        merged.append(cur)
        bi += 1

    out_path = Path(__file__).resolve().parent / "siseol-parsed.json"
    out_path.write_text(json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8")
    print(out_path)


if __name__ == "__main__":
    main()
